# # pip install streamlit azure-ai-projects azure-identity pandas python-dotenv openpyxl

# import streamlit as st
# import pandas as pd
# from azure.identity import DefaultAzureCredential
# from azure.ai.projects import AIProjectClient
# from dotenv import load_dotenv
# import os

# from openpyxl import Workbook
# from openpyxl.styles import Alignment, Font
# from openpyxl.utils import get_column_letter
# from io import BytesIO

# load_dotenv()

# # -------------------------
# # Azure Config
# # -------------------------
# MY_ENDPOINT = "https://airports-ai-team-resource.services.ai.azure.com/api/projects/airports-ai-team"
# AGENT_NAME = "airports-websearch"
# AGENT_VERSION = "9"

# @st.cache_resource
# def get_client():
#     project_client = AIProjectClient(
#         endpoint=MY_ENDPOINT,
#         credential=DefaultAzureCredential(),
#     )
#     return project_client.get_openai_client()

# openai_client = get_client()

# # -------------------------
# # Streamlit UI
# # -------------------------
# st.set_page_config(page_title="ESG Report Generator", layout="wide")
# st.title(" ESG Data Intelligence Agent")

# # -------------------------
# # Session State
# # -------------------------
# if "companies" not in st.session_state:
#     st.session_state.companies = [{"name": "", "website": ""}]

# st.subheader(" Enter Company Details")

# # -------------------------
# # Dynamic Inputs
# # -------------------------
# for i, company in enumerate(st.session_state.companies):

#     col1, col2, col3 = st.columns([4, 4, 1])

#     with col1:
#         st.session_state.companies[i]["name"] = st.text_input(
#             f"Company Name {i+1}",
#             value=company["name"],
#             key=f"name_{i}"
#         )

#     with col2:
#         st.session_state.companies[i]["website"] = st.text_input(
#             f"Website {i+1}",
#             value=company["website"],
#             key=f"web_{i}"
#         )

#     with col3:
#         if st.button("Delete", key=f"remove_{i}"):
#             st.session_state.companies.pop(i)
#             st.rerun()

# # Add Company
# if st.button(" Add Company"):
#     st.session_state.companies.append({"name": "", "website": ""})
#     st.rerun()

# # -------------------------
# # Fetch ESG Report
# # -------------------------
# if st.button(" Fetch ESG Report"):

#     valid_companies = [
#         c for c in st.session_state.companies if c["name"].strip()
#     ]

#     if not valid_companies:
#         st.warning(" Please enter at least one company")
#         st.stop()

#     with st.spinner("Fetching ESG data... "):

#         try:
#             all_rows = []

#             # -------------------------
#             # API Loop
#             # -------------------------
#             for company in valid_companies:

#                 company_name = company["name"]
#                 company_website = company["website"]

#                 user_prompt = f"""
# Company Name: {company_name}
# Official Website: {company_website if company_website else "Not provided"}

# Return ONLY a clean table using '|' delimiter.

# Columns:
# Topic | Detailed Information Retrieved | Reference Link

# make sure that the reference link must be reference web page or report. and accessible
# """

#                 response = openai_client.responses.create(
#                     input=[{"role": "user", "content": user_prompt}],
#                     extra_body={
#                         "agent_reference": {
#                             "name": AGENT_NAME,
#                             "version": AGENT_VERSION,
#                             "type": "agent_reference"
#                         }
#                     },
#                 )

#                 output = response.output_text.strip()
#                 lines = output.split("\n")

#                 for line in lines:
#                     if "|" in line:
#                         parts = [col.strip() for col in line.split("|")]

#                         if len(parts) >= 3 and parts[0].lower() != "topic":
#                             all_rows.append([
#                                 company_name,
#                                 parts[0],
#                                 parts[1],
#                                 parts[2]
#                             ])

#             # -------------------------
#             # DataFrame
#             # -------------------------
#             if all_rows:

#                 df = pd.DataFrame(
#                     all_rows,
#                     columns=[
#                         "Company Name",
#                         "Topic",
#                         "Detailed Information Retrieved",
#                         "Reference Link"
#                     ]
#                 )

#                 # -------------------------
#                 # Clickable links for UI
#                 # -------------------------
#                 def make_clickable(link):
#                     if isinstance(link, str) and link.startswith("http"):
#                         return f'<a href="{link}" target="_blank">{link}</a>'
#                     return link

#                 df_display = df.copy()
#                 df_display["Reference Link"] = df_display["Reference Link"].apply(make_clickable)

#                 html_table = df_display.to_html(escape=False, index=False)

#                 st.success(" ESG Report Generated")
#                 st.subheader(" ESG Report")

#                 # -------------------------
#                 # Styling
#                 # -------------------------
#                 st.markdown("""
#                     <style>
#                     table {
#                         width: 100%;
#                         border-collapse: collapse;
#                         table-layout: fixed;
#                     }
#                     th {
#                         background-color: #262730;
#                         color: white;
#                         padding: 12px;
#                         text-align: left;
#                     }
#                     td {
#                         padding: 12px;
#                         border-bottom: 1px solid #ddd;
#                         vertical-align: top;
#                         word-wrap: break-word;
#                         white-space: normal !important;
#                         font-size: 14px;
#                         line-height: 1.5;
#                     }
#                     tr:hover {
#                         background-color: #f5f5f5;
#                     }

#                     th:nth-child(1), td:nth-child(1) { width: 15%; }
#                     th:nth-child(2), td:nth-child(2) { width: 20%; }
#                     th:nth-child(3), td:nth-child(3) { width: 45%; }
#                     th:nth-child(4), td:nth-child(4) { width: 20%; }
#                     </style>
#                 """, unsafe_allow_html=True)

#                 st.markdown(html_table, unsafe_allow_html=True)

#                 # -------------------------
#                 # Excel Generation
#                 # -------------------------
#                 output_excel = BytesIO()
#                 wb = Workbook()
#                 wb.remove(wb.active)

#                 for company in df["Company Name"].unique():

#                     company_df = df[df["Company Name"] == company]

#                     ws = wb.create_sheet(title=company[:30])

#                     # Headers
#                     headers = list(company_df.columns)
#                     ws.append(headers)

#                     for col_num, col_name in enumerate(headers, 1):
#                         cell = ws.cell(row=1, column=col_num)
#                         cell.font = Font(bold=True)
#                         cell.alignment = Alignment(wrap_text=True)

#                     # Data
#                     for _, row in company_df.iterrows():
#                         ws.append(list(row))

#                     # Style cells
#                     for row in ws.iter_rows():
#                         for cell in row:
#                             cell.alignment = Alignment(wrap_text=True, vertical="top")

#                     # Column widths
#                     col_widths = [20, 25, 60, 40]
#                     for i, width in enumerate(col_widths, 1):
#                         ws.column_dimensions[get_column_letter(i)].width = width

#                 wb.save(output_excel)
#                 output_excel.seek(0)

#                 # -------------------------
#                 # Download Excel
#                 # -------------------------
#                 st.download_button(
#                     label=" Download Excel (Formatted + Multi-Sheet)",
#                     data=output_excel,
#                     file_name="ESG_Multi_Company_Report.xlsx",
#                     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#                 )

#             else:
#                 st.warning(" Could not structure table properly.")

#         except Exception as e:
#             st.error(f" Error: {str(e)}")



# pip install streamlit azure-ai-projects azure-identity pandas python-dotenv openpyxl

import streamlit as st
import pandas as pd
from azure.identity import DefaultAzureCredential
from azure.ai.projects import AIProjectClient
from dotenv import load_dotenv
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from io import BytesIO

load_dotenv()

# -------------------------
# Azure Config
# -------------------------
MY_ENDPOINT = "https://airports-ai-team-resource.services.ai.azure.com/api/projects/airports-ai-team"
AGENT_NAME = "airports-websearch"
AGENT_VERSION = "9"

# -------------------------
# Azure Client
# -------------------------
@st.cache_resource
def get_client():
    project_client = AIProjectClient(
        endpoint=MY_ENDPOINT,
        credential=DefaultAzureCredential(),
    )
    return project_client.get_openai_client()

openai_client = get_client()

# -------------------------
# Fetch ESG Data (API)
# -------------------------
def fetch_esg_data(companies):
    all_rows = []

    for company in companies:
        company_name = company["name"]
        company_website = company["website"]

        user_prompt = f"""
Company Name: {company_name}
Official Website: {company_website if company_website else "Not provided"}

Return ONLY a clean table using '|' delimiter.

Columns:
Topic | Detailed Information Retrieved | Reference Link

make sure that the reference link must be reference web page or report. and accessible
"""

        response = openai_client.responses.create(
            input=[{"role": "user", "content": user_prompt}],
            extra_body={
                "agent_reference": {
                    "name": AGENT_NAME,
                    "version": AGENT_VERSION,
                    "type": "agent_reference"
                }
            },
        )

        output = response.output_text.strip()
        lines = output.split("\n")

        for line in lines:
            if "|" in line:
                parts = [col.strip() for col in line.split("|")]

                if len(parts) >= 3 and parts[0].lower() != "topic":
                    all_rows.append([
                        company_name,
                        parts[0],
                        parts[1],
                        parts[2]
                    ])

    if all_rows:
        return pd.DataFrame(
            all_rows,
            columns=[
                "Company Name",
                "Topic",
                "Detailed Information Retrieved",
                "Reference Link"
            ]
        )

    return None


# -------------------------
# Excel Generator
# -------------------------
def generate_excel(df):
    output_excel = BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    for company in df["Company Name"].unique():

        company_df = df[df["Company Name"] == company]
        ws = wb.create_sheet(title=company[:30])

        headers = list(company_df.columns)
        ws.append(headers)

        # Header styling
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True)

        # Data
        for _, row in company_df.iterrows():
            ws.append(list(row))

        # Cell styling
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Column width
        col_widths = [20, 25, 60, 40]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    wb.save(output_excel)
    output_excel.seek(0)

    return output_excel


# -------------------------
# Display Table
# -------------------------
def display_table(df):

    def make_clickable(link):
        if isinstance(link, str) and link.startswith("http"):
            return f'<a href="{link}" target="_blank">{link}</a>'
        return link

    df_display = df.copy()
    df_display["Reference Link"] = df_display["Reference Link"].apply(make_clickable)

    html_table = df_display.to_html(escape=False, index=False)

    st.subheader(" ESG Report")

    st.markdown("""
        <style>
        table {
            width: 100%;
            border-collapse: collapse;
            table-layout: fixed;
        }
        th {
            background-color: #262730;
            color: white;
            padding: 12px;
            text-align: left;
        }
        td {
            padding: 12px;
            border-bottom: 1px solid #ddd;
            vertical-align: top;
            word-wrap: break-word;
            white-space: normal !important;
            font-size: 14px;
            line-height: 1.5;
        }
        tr:hover {
            background-color: #f5f5f5;
        }

        th:nth-child(1), td:nth-child(1) { width: 15%; }
        th:nth-child(2), td:nth-child(2) { width: 20%; }
        th:nth-child(3), td:nth-child(3) { width: 45%; }
        th:nth-child(4), td:nth-child(4) { width: 20%; }
        </style>
    """, unsafe_allow_html=True)

    st.markdown(html_table, unsafe_allow_html=True)


# -------------------------
# Streamlit UI
# -------------------------
st.set_page_config(page_title="ESG Report Generator", layout="wide")
st.title(" ESG Data Intelligence Agent")

if "companies" not in st.session_state:
    st.session_state.companies = [{"name": "", "website": ""}]

st.subheader(" Enter Company Details")

# Dynamic Inputs
for i, company in enumerate(st.session_state.companies):

    col1, col2, col3 = st.columns([4, 4, 1])

    with col1:
        st.session_state.companies[i]["name"] = st.text_input(
            f"Company Name {i+1}",
            value=company["name"],
            key=f"name_{i}"
        )

    with col2:
        st.session_state.companies[i]["website"] = st.text_input(
            f"Website {i+1}",
            value=company["website"],
            key=f"web_{i}"
        )

    with col3:
        if st.button("Delete", key=f"remove_{i}"):
            st.session_state.companies.pop(i)
            st.rerun()

if st.button(" Add Company"):
    st.session_state.companies.append({"name": "", "website": ""})
    st.rerun()

# -------------------------
# Fetch Button
# -------------------------
if st.button(" Fetch ESG Report"):

    valid_companies = [c for c in st.session_state.companies if c["name"].strip()]

    if not valid_companies:
        st.warning(" Please enter at least one company")
    else:
        with st.spinner("Fetching ESG data..."):
            df = fetch_esg_data(valid_companies)

            if df is not None:
                st.session_state["df"] = df
            else:
                st.warning(" Could not structure table properly.")

# -------------------------
# Show Results (NO REFRESH ISSUE)
# -------------------------
if "df" in st.session_state:

    st.success(" ESG Report Generated")

    df = st.session_state["df"]

    display_table(df)

    excel_file = generate_excel(df)

    st.download_button(
        label=" Download Excel",
        data=excel_file,
        file_name="ESG_Multi_Company_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )